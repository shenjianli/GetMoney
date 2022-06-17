# -*- coding: utf-8 -*-
import requests
import openpyxl
import json

from fund_pipelines.FundLineShow import FundLineShow


class FundPipeLine(object):
    def __init__(self):
        self.tag = "FundPipeLine"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0'
        }
        self.url = "https://danjuanapp.com/djapi/v3/filter/fund?type={}&order_by=1w&size=10&page=1"

        # 基金类型
        self.dict_type = {"股票型": 1, "混合型": 3, "债券型": 2, "指数型": 5, "QDII型": 11}
        # 时间 order_by
        self.dict_time = {'近一周': '1w', '近一月': '1m', '近三月': '3m', '近六月': '6m', '近1年': '1y',
                          '近2年': '2y', '近3年': '3y', '近5年': '5y'}

    def log(self, msg):
        print(self.tag, msg)

    def everyone(self, code):
        # code="008189"
        url = "https://danjuanapp.com/djapi/fund/nav/history/" + str(code) + "?page=1&size=190"
        res = requests.get(url, headers=self.headers)
        res.encoding = 'utf-8'
        s = json.loads(res.text)
        s = s['data']['items']
        date = []
        percentage = []
        for i in range(0, len(s)):
            print(s[i]['date'] + ":" + s[i]['percentage'])

            date.append(s[i]['date'])
            percentage.append(s[i]['percentage'])
        datelist = []
        history = []

        history.append(float(percentage[len(percentage) - 1]))
        datelist.append(date[len(date) - 1])
        for i in range(len(percentage) - 2, -1, -1):
            history.append(float(history[len(percentage) - i - 2]) + float(percentage[i]))
            datelist.append(date[i])
        return datelist, history

    def getlist(self):
        for key in self.dict_type:

            outwb = openpyxl.Workbook()
            outws = outwb.create_sheet(index=0)
            outws.cell(row=1, column=1, value="日期")

            column = 2
            url = "https://danjuanapp.com/djapi/v3/filter/fund?type=" + str(
                self.dict_type[key]) + "&order_by=1y&size=20&page=1"
            res = requests.get(url, headers=self.headers)
            res.encoding = 'utf-8'
            s = json.loads(res.text)
            s = s['data']['items']
            for i in range(0, len(s)):
                print(s[i]['fd_name'] + ":" + s[i]['fd_code'])

                outws.cell(row=1, column=column, value=str(s[i]['fd_name']))

                datelist, daydetail = self.everyone(s[i]['fd_code'])

                count = 2
                ###写入日期
                if i == 0:
                    for k in range(0, len(datelist)):
                        outws.cell(row=count, column=1, value=str(datelist[k]))
                        count = count + 1

                count = 2
                ##写入涨幅
                for k in range(0, len(daydetail)):
                    outws.cell(row=count, column=column, value=str(str(round(float(daydetail[k]), 2))))
                    count = count + 1

                # value.append(s[i]['fd_code'])
                column = column + 1
            print("-------------")

            outwb.save(str(key) + "近一年日涨幅.xlsx")  # 保存


if __name__ == '__main__':
    # fund_info = FundPipeLine()
    # fund_info.getlist()
    fund_line_show = FundLineShow()
    fund_line_show.show("债券型近一年日涨幅")
